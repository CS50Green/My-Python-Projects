import openpyxl
import openpyxl.styles
import notifypy
import time
import os


def main():
    # Checking is "Tasks.xlsx" file exists or not
    # :os.path.isfile(): will return True if file exists otherwise False
    # If file doesn't exist then call create_workbook() function
    exists = os.path.isfile("Tasks.xlsx")
    if exists is False:
        create_workbook()

    # Loading "Tasks.xlsx" workbook
    wb = openpyxl.load_workbook("Tasks.xlsx")
    # :wb.active: Will grab the sheet that we are currently working in
    ws = wb.active

    # Looping through every row in "Tasks.xlsx" and calling remind() function
    for row in ws.iter_rows(min_row=2, values_only=True):
        remind(row[0], row[1:])


# This function will create a Workbook and write header
def create_workbook():
    # :openpyxl.Workbook(): Will create a new Workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    # Adding header to first columns
    ws["A1"] = "Tasks"
    ws["B1"] = "First Remind Time"
    ws["C1"] = "Second Remind Time"
    ws["D1"] = "Third Remind Time"
    ws["E1"] = "Fourth Remind Time"
    ws["F1"] = "Fifth Remind Time"

    # Adjusting width of columns
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 30

    # Creating a font style
    font_style = openpyxl.styles.Font(size=15, bold=True)

    # Putting font_style in columns
    ws["A1"].font = font_style
    ws["B1"].font = font_style
    ws["C1"].font = font_style
    ws["D1"].font = font_style
    ws["E1"].font = font_style
    ws["F1"].font = font_style

    # Centering the text
    ws["A1"].alignment = openpyxl.styles.Alignment(horizontal="center")
    ws["B1"].alignment = openpyxl.styles.Alignment(horizontal="center")
    ws["C1"].alignment = openpyxl.styles.Alignment(horizontal="center")
    ws["D1"].alignment = openpyxl.styles.Alignment(horizontal="center")
    ws["E1"].alignment = openpyxl.styles.Alignment(horizontal="center")
    ws["F1"].alignment = openpyxl.styles.Alignment(horizontal="center")

    # Saving workbook
    wb.save("Tasks.xlsx")


# This function will send notifications
def remind(message, times_list):
    notification = notifypy.Notify(default_notification_application_name="Python Application (reminder.py)")
    notification.title = "Reminder"
    notification.message = message

    # Adding icons to our notifications
    if message == "Notion Entry Time":
        notification.icon = "Icons/Notion_Icon.png"
    elif message == "Time To Workout":
        notification.icon = "Icons/Workout_Icon.png"
    elif message == "Time To Pray Namaz":
        notification.icon = "Icons/Praying_Namaz_Icon.jpg"

    # :times_list: Contains times that tells when to send notifications
    # If current time is equal to a time in times_list then notification will be sent
    for t in times_list:
        if t == time.strftime("%I:%M %p"):
            notification.send()


if __name__ == "__main__":
    main()
